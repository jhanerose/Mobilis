#!/usr/bin/env python3
import sys
import csv
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import TableStyle, Table
from db_client import execute_query

def export_csv(data, output_file):
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Customer ID', 'Name', 'Email', 'Phone', 'Role', 'Bookings', 'Spent', 'License Number', 'License Expiry', 'Address'])
        for row in data:
            writer.writerow([
                row['user_id'],
                row['name'],
                row['email'],
                row['phone'],
                row['role'],
                row['bookings'],
                row['spent'],
                row['license_number'],
                row['license_expiry'],
                row['address']
            ])

def export_xlsx(data, output_file):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define styles - using Mobilis brand colors
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="16986D", end_color="16986D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="DBE3DE"),
        right=Side(style='thin', color="DBE3DE"),
        top=Side(style='thin', color="DBE3DE"),
        bottom=Side(style='thin', color="DBE3DE")
    )
    
    # Add title
    ws['A1'] = "Mobilis Vehicle Rental - Customers Report"
    ws['A1'].font = Font(bold=True, size=16, color="16986D")
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add generated date
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(italic=True, size=10, color="6A7577")
    ws.merge_cells('A2:J2')
    
    # Add headers
    headers = ['Customer ID', 'Name', 'Email', 'Phone', 'Role', 'Bookings', 'Spent', 'License Number', 'License Expiry', 'Address']
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    for row in data:
        ws.append([
            row['user_id'],
            row['name'],
            row['email'],
            row['phone'],
            row['role'],
            row['bookings'],
            row['spent'],
            row['license_number'],
            row['license_expiry'],
            row['address']
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.column == 6:  # Bookings column
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column == 7:  # Spent column
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '₱#,##0.00'
    
    # Auto-adjust column widths
    column_widths = [12, 25, 30, 15, 12, 10, 12, 18, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A4'
    
    wb.save(output_file)

def export_pdf(data, output_file):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime
    
    doc = SimpleDocTemplate(output_file, pagesize=landscape(letter), topMargin=72, bottomMargin=36, leftMargin=36, rightMargin=36)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Custom styles - using Mobilis brand colors
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#16986D'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6A7577'),
        alignment=TA_CENTER,
        spaceAfter=24
    )
    
    # Add title
    elements.append(Paragraph("Mobilis Vehicle Rental - Customers Report", title_style))
    
    # Add generated date
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    elements.append(Spacer(1, 12))
    
    headers = ['Customer ID', 'Name', 'Email', 'Phone', 'Role', 'Bookings', 'Spent', 'License Number', 'License Expiry', 'Address']
    table_data = [headers]
    
    for row in data:
        table_data.append([
            str(row['user_id']),
            row['name'],
            row['email'],
            row['phone'],
            row['role'],
            str(row['bookings']),
            f"₱{row['spent']:.2f}",
            row['license_number'] or '',
            row['license_expiry'] or '',
            row['address'] or ''
        ])
    
    table = Table(table_data, colWidths=[50, 100, 120, 60, 50, 50, 60, 70, 70, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16986D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6F9F7')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DBE3DE')),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)

def main():
    format_type = sys.argv[1] if len(sys.argv) > 1 else 'csv'
    output_file = sys.argv[2] if len(sys.argv) > 2 else f'customers_export.{format_type}'

    query = """
        SELECT 
            u.user_id,
            CONCAT(u.first_name, ' ', u.last_name) as name,
            u.email,
            u.phone,
            u.role,
            COUNT(r.rental_id) as bookings,
            COALESCE(SUM(i.total_amount), 0) as spent,
            u.license_number,
            u.license_expiry,
            u.address
        FROM User u
        LEFT JOIN Rental r ON u.user_id = r.user_id
        LEFT JOIN Invoice i ON r.rental_id = i.rental_id
        GROUP BY u.user_id, u.first_name, u.last_name, u.email, u.phone, u.role, u.license_number, u.license_expiry, u.address
        ORDER BY spent DESC
        LIMIT 1000
    """

    customers = execute_query(query)

    # Export based on format
    if format_type == 'xlsx':
        export_xlsx(customers, output_file)
    elif format_type == 'pdf':
        export_pdf(customers, output_file)
    else:
        export_csv(customers, output_file)

    print(f"Exported {len(customers)} customers to {output_file}")

if __name__ == "__main__":
    main()
