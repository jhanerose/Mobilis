#!/usr/bin/env python3
import sys
import csv
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import TableStyle, Table
from db_client import execute_query

def export_csv(data, output_file):
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Invoice ID', 'Rental ID', 'Customer', 'Vehicle', 'Base Amount', 'Late Fee', 'Damage Fee', 'Total', 'Payment Status', 'Payment Method', 'Issued At'])
        for row in data:
            writer.writerow([
                row['invoice_id'],
                row['rental_id'],
                row['customer'],
                row['vehicle'],
                row['base_amount'],
                row['late_fee'],
                row['damage_fee'],
                row['total_amount'],
                row['payment_status'],
                row['payment_method'],
                row['issued_at']
            ])

def export_xlsx(data, output_file):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Define styles - using Mobilis brand colors
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="16986D", end_color="16986D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="DBE3DE"),
        right=Side(style='thin', color="DBE3DE"),
        top=Side(style='thin', color="DBE3DE"),
        bottom=Side(style='thin', color="DBE3DE")
    )
    
    # Add title
    ws['A1'] = "Mobilis Vehicle Rental - Payments Report"
    ws['A1'].font = Font(bold=True, size=16, color="16986D")
    ws.merge_cells('A1:K1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add generated date
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(italic=True, size=10, color="6A7577")
    ws.merge_cells('A2:K2')
    
    # Add headers
    headers = ['Invoice ID', 'Rental ID', 'Customer', 'Vehicle', 'Base Amount', 'Late Fee', 'Damage Fee', 'Total', 'Payment Status', 'Payment Method', 'Issued At']
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    for row in data:
        ws.append([
            row['invoice_id'],
            row['rental_id'],
            row['customer'],
            row['vehicle'],
            row['base_amount'],
            row['late_fee'],
            row['damage_fee'],
            row['total_amount'],
            row['payment_status'],
            row['payment_method'],
            row['issued_at']
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.column in [5, 6, 7, 8]:  # Amount columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '₱#,##0.00'
    
    # Auto-adjust column widths
    column_widths = [12, 12, 25, 25, 12, 10, 10, 12, 15, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A4'
    
    wb.save(output_file)

def export_pdf(data, output_file):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime
    
    doc = SimpleDocTemplate(output_file, pagesize=landscape(letter), topMargin=72, bottomMargin=36, leftMargin=36, rightMargin=36)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Custom styles - using Mobilis brand colors
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#16986D'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6A7577'),
        alignment=TA_CENTER,
        spaceAfter=24
    )
    
    # Add title
    elements.append(Paragraph("Mobilis Vehicle Rental - Payments Report", title_style))
    
    # Add generated date
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    elements.append(Spacer(1, 12))
    
    headers = ['Invoice ID', 'Rental ID', 'Customer', 'Vehicle', 'Base Amount', 'Late Fee', 'Damage Fee', 'Total', 'Payment Status', 'Payment Method', 'Issued At']
    table_data = [headers]
    
    for row in data:
        table_data.append([
            str(row['invoice_id']),
            str(row['rental_id']),
            row['customer'],
            row['vehicle'],
            f"₱{row['base_amount']:.2f}",
            f"₱{row['late_fee']:.2f}",
            f"₱{row['damage_fee']:.2f}",
            f"₱{row['total_amount']:.2f}",
            row['payment_status'],
            row['payment_method'],
            row['issued_at']
        ])
    
    table = Table(table_data, colWidths=[50, 50, 100, 100, 60, 50, 50, 60, 70, 70, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16986D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6F9F7')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DBE3DE')),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)

def main():
    status = sys.argv[1] if len(sys.argv) > 1 else ''
    from_date = sys.argv[2] if len(sys.argv) > 2 else ''
    to_date = sys.argv[3] if len(sys.argv) > 3 else ''
    search = sys.argv[4] if len(sys.argv) > 4 else ''
    format_type = sys.argv[5] if len(sys.argv) > 5 else 'csv'
    output_file = sys.argv[6] if len(sys.argv) > 6 else f'payments_export.{format_type}'

    query = """
        SELECT 
            i.invoice_id,
            r.rental_id,
            CONCAT(u.first_name, ' ', u.last_name) as customer,
            CONCAT(v.brand, ' ', v.model) as vehicle,
            i.base_amount,
            i.late_fee,
            i.damage_fee,
            i.total_amount,
            i.payment_status,
            i.payment_method,
            i.issued_at
        FROM Invoice i
        JOIN Rental r ON i.rental_id = r.rental_id
        JOIN User u ON r.user_id = u.user_id
        JOIN Vehicle v ON r.vehicle_id = v.vehicle_id
        WHERE 1=1
    """
    params = []

    if status:
        query += " AND i.payment_status = %s"
        params.append(status)
    if from_date:
        query += " AND DATE(i.issued_at) >= %s"
        params.append(from_date)
    if to_date:
        query += " AND DATE(i.issued_at) <= %s"
        params.append(to_date)

    query += " ORDER BY i.issued_at DESC LIMIT 2000"

    payments = execute_query(query, tuple(params))

    # Apply search filter
    if search:
        payments = [p for p in payments if search.lower() in str(p).lower()]

    # Export based on format
    if format_type == 'xlsx':
        export_xlsx(payments, output_file)
    elif format_type == 'pdf':
        export_pdf(payments, output_file)
    else:
        export_csv(payments, output_file)

    print(f"Exported {len(payments)} payments to {output_file}")

if __name__ == "__main__":
    main()
